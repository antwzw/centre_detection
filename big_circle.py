# 相机参数的存放位置, 要求为 xlsx 格式
CAMERA_PARA_PATH = '/Users/zhengbaoqin/Desktop/bigcircle/Code1109/4格_亚克力板_新.xlsx'
# 输入图片位置
IMG_PATH = '/Users/zhengbaoqin/Desktop/bigcircle/Code1109/亚克力板'
# 照片数量
IMG_NUM = 8
# 结果输出位置, 要求为 xlsx 格式
OUTPUT = '/Users/zhengbaoqin/Desktop/bigcircle/Code1109/result.xlsx'
# 棋盘格的宽度、高度
WIDTH = 18
HEIGHT = 18
# 透视变换后的高度相当于多少个棋盘格
SIZE = 40
# 是否为部分棋盘格
USE_PART_CHESSBOARD = True
# 部分棋盘格的尺寸参数, 仅在使用了部分棋盘格的情况下需要填写
# 棋盘格尺寸: 5 个参数, 从左到右依次为 宽度、高度、左侧宽度、右侧宽度、上方高度 (尺寸为方格数量)
CHESSBOARD_SIZE = (32, 32, 6, 6, 6)
# 不要管这个参数
CENTER = (0.0, 0.0)
# 大圆圆形结果选取时允许的偏差, 不建议修改这个参数
ALLOWED_DISTANCE = 8.0


import cv2
import numpy as np
import openpyxl
from openpyxl import Workbook
import os

import separate_cb
separate_cb.CHESSBOARD_SIZE = CHESSBOARD_SIZE
from separate_cb import init, getFourCorners
init()


if (__name__ == '__main__'):
    separate_cb.CHESSBOARD_SIZE = CHESSBOARD_SIZE


# HED 要用的 Class
class CropLayer(object):
    def __init__(self, params, blobs):
        # 初始化剪切区域开始和结束点的坐标
        self.xstart = 0
        self.ystart = 0
        self.xend = 0
        self.yend = 0

    # 计算输入图像的体积
    def getMemoryShapes(self, inputs):
        (inputShape, targetShape) = (inputs[0], inputs[1])
        (batchSize, numChannels) = (inputShape[0], inputShape[1])
        (H, W) = (targetShape[2], targetShape[3])

        # 计算开始和结束剪切坐标的值
        self.xstart = int((inputShape[3] - targetShape[3]) // 2)
        self.ystart = int((inputShape[2] - targetShape[2]) // 2)
        self.xend = self.xstart + W
        self.yend = self.ystart + H

        # 返回体积，接下来进行实际裁剪
        return [[batchSize, numChannels, H, W]]

    def forward(self, inputs):
        return [inputs[0][:, :, self.ystart:self.yend, self.xstart:self.xend]]


# 定义 HED 函数
def HED(image, net, downW=800, downH=800):
    (H, W) = image.shape[:2]
    image = cv2.resize(image, (downW, downH))
    # 根据输入图像为全面的嵌套边缘检测器（Holistically-Nested Edge Detector）构建一个输出blob
    blob = cv2.dnn.blobFromImage(image, scalefactor=1.0, size=(800, 800),
                                 mean=(104.00698793, 116.66876762, 122.67891434),
                                 # mean=(0, 0, 0),
                                 swapRB=False, crop=True)

    # # 设置blob作为网络的输入并执行算法以计算边缘图
    print("[INFO] performing holistically-nested edge detection...")
    net.setInput(blob)
    hed = net.forward()
    # 调整输出为原始图像尺寸的大小
    hed = cv2.resize(hed[0, 0], (W, H))
    # 将图像像素缩回到范围[0,255]并确保类型为“UINT8”
    hed = (255 * hed).astype("uint8")
    return hed


# 确保这些代码只执行一次
try:
    _tmp2 = _tmp1
except:
    _tmp1 = 0
    protoPath = "/Users/zhengbaoqin/Desktop/project/code/deploy.prototxt"
    modelPath = "/Users/zhengbaoqin/Desktop/project/code/hed_pretrained_bsds.caffemodel"
    net = cv2.dnn.readNetFromCaffe(protoPath, modelPath)
    cv2.dnn_registerLayer("Crop", CropLayer)


# 先声明需要使用的全局变量
mtx = None
dist = None
newcameramtx = None


def getPath(n: int):
    n = str(n)
    if (len(n) == 1):
        n = '0' + n
    return f'{IMG_PATH}/{n}.jpg'


# Excel 部分
ALPHAS = '.ABCDEFGHIJKLMNOPQRSTUVWXYZ'
NUMS = {}
for i in range(1, 27):
    NUMS[ALPHAS[i]] = i


def numToAlpha(n):
    s = ''
    while (n > 0):
        t = n % 26
        if (t == 0):
            t = 26
        s = ALPHAS[t] + s
        n = (n - t) // 26
    return s


def alphaToNum(a):
    l = len(a)
    sum = 0
    for i in range(0, l):
        sum += 26**(l - i - 1) * NUMS[a[i]]
    return sum


def cellToTuple(cell: str):
    for i in range(0, len(cell)):
        try:
            a = NUMS[cell[i]]
        except:
            n = i
            break
    return (alphaToNum(cell[:n]), int(cell[n:]))


def tupleToCell(t):
    return numToAlpha(t[0]) + str(t[1])


def read_cell(sheet, i, j):
    return sheet[tupleToCell((i, j))].value


def write_cell(sheet, i, j, value):
    sheet[tupleToCell((i, j))].value = value


# 从 Excel 读取相机参数的函数
def read_camera_matrix():
    mtx = np.zeros((3, 3), np.float64)
    newcameramtx = np.zeros((3, 3), np.float64)
    dist = np.zeros((1, 5), np.float64)
    wb = openpyxl.load_workbook(CAMERA_PARA_PATH)
    ws = wb.active
    mtx[0][0] = read_cell(ws, 1, 1)
    mtx[0][1] = read_cell(ws, 2, 1)
    mtx[0][2] = read_cell(ws, 3, 1)
    mtx[1][0] = read_cell(ws, 1, 2)
    mtx[1][1] = read_cell(ws, 2, 2)
    mtx[1][2] = read_cell(ws, 3, 2)
    mtx[2][0] = read_cell(ws, 1, 3)
    mtx[2][1] = read_cell(ws, 2, 3)
    mtx[2][2] = read_cell(ws, 3, 3)
    newcameramtx[0][0] = read_cell(ws, 1, 5)
    newcameramtx[0][1] = read_cell(ws, 2, 5)
    newcameramtx[0][2] = read_cell(ws, 3, 5)
    newcameramtx[1][0] = read_cell(ws, 1, 6)
    newcameramtx[1][1] = read_cell(ws, 2, 6)
    newcameramtx[1][2] = read_cell(ws, 3, 6)
    newcameramtx[2][0] = read_cell(ws, 1, 7)
    newcameramtx[2][1] = read_cell(ws, 2, 7)
    newcameramtx[2][2] = read_cell(ws, 3, 7)
    dist[0][0] = read_cell(ws, 1, 9)
    dist[0][1] = read_cell(ws, 2, 9)
    dist[0][2] = read_cell(ws, 3, 9)
    dist[0][3] = read_cell(ws, 4, 9)
    dist[0][4] = read_cell(ws, 5, 9)
    return (mtx, dist, newcameramtx)


# 读取所有图片
def read_all():
    all = []
    for i in range(1, IMG_NUM + 1):
        all.append(cv2.imread(getPath(i)))
    return all


# 对所有图片去除畸变
def undistort_all(all_img: list):
    result = []
    for img in all_img:
        dst = cv2.undistort(img, mtx, dist, None, newcameramtx)
        result.append(dst)
    return result


# 在棋盘格所有格点中挑选出角上的, 被 transform_all 调用
def getCorner2(points: np.ndarray, Type: str):
    # type  upleft, upright, downleft, downright
    dic = {
        'upleft': [1, 1],
        'upright': [-1, 1],
        'downleft': [1, -1],
        'downright': [-1, -1]
    }
    mode = dic[Type]
    sum = []
    for p in points:
        sum.append(p[0][0] * mode[0] + p[0][1] * mode[1])
    loc = sum.index(min(sum))
    return points[loc][0]


# 对所有图片透视变换
def transform_all(all_img: list):
    global CENTER
    result = []
    if (USE_PART_CHESSBOARD):
        for img in all_img:
            cv2.imwrite('./._tmp_.png', img)
            fourCorners = getFourCorners('./._tmp_.png')
            os.remove('./._tmp_.png')
            if (fourCorners[0] == False):
                print(fourCorners[1])
                continue
            upleft = fourCorners[1][0]
            upright = fourCorners[1][1]
            downleft = fourCorners[1][2]
            downright = fourCorners[1][3]
            h, w = img.shape[:2]
            cap = max(h, w)
            cap = h
            CENTER = (float(0.5 * cap), float(0.5 * cap))
            x_size = SIZE
            objPoints = np.float32([[x_size / 2 - WIDTH / 2 + 0.5, SIZE / 2 - HEIGHT / 2 + 0.5], [x_size / 2 + WIDTH / 2 - 0.5, SIZE / 2 - HEIGHT / 2 + 0.5], [x_size / 2 - WIDTH / 2 + 0.5, SIZE / 2 + HEIGHT / 2 - 0.5], [x_size / 2 + WIDTH / 2 - 0.5, SIZE / 2 + HEIGHT / 2 - 0.5]]) * cap / SIZE
            imgPoints = np.float32([upleft, upright, downleft, downright])
            m = cv2.getPerspectiveTransform(imgPoints, objPoints)
            new = cv2.warpPerspective(img, m, (cap, cap))
            result.append(new)
        return result
    i = 0
    for img in all_img:
        i += 1
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        h, w = img.shape[:2]
        cap = h
        CENTER = (float(0.5 * cap), float(0.5 * cap))
        x_size = SIZE
        objPoints = np.float32(
            [[x_size / 2 - WIDTH / 2 + 0.5, SIZE / 2 - HEIGHT / 2 + 0.5], [x_size / 2 + WIDTH / 2 - 0.5, SIZE / 2 - HEIGHT / 2 + 0.5], [x_size / 2 - WIDTH / 2 + 0.5, SIZE / 2 + HEIGHT / 2 - 0.5], [x_size / 2 + WIDTH / 2 - 0.5, SIZE / 2 + HEIGHT / 2 - 0.5]]) * cap / SIZE
        corners = cv2.findChessboardCorners(gray, (WIDTH, HEIGHT), None)
        if (corners[0] == False):
            print(f'第{i}张图片透视变换是未成功识别棋盘格, 请检查原因')
            continue
        corners = corners[1]
        corners = cv2.cornerSubPix(gray, corners, (5, 4), (-1, -1),
                                   (cv2.TERM_CRITERIA_MAX_ITER | cv2.TERM_CRITERIA_EPS, 30, 0.001))
        upleft = getCorner2(corners, 'upleft')
        upright = getCorner2(corners, 'upright')
        downleft = getCorner2(corners, 'downleft')
        downright = getCorner2(corners, 'downright')
        imgPoints = np.float32([upleft, upright, downleft, downright])
        m = cv2.getPerspectiveTransform(imgPoints, objPoints)
        new = cv2.warpPerspective(img, m, (cap, cap))
        result.append(new)
    return result


# 对所有图片进行 HED
def hed_all(all_img: list):
    result = []
    for img in all_img:
        size = img.shape
        ratio = 1
        ratiox = int(size[1] / ratio)
        ratioy = int(size[0] / ratio)
        img = cv2.resize(img, (ratiox, ratioy))
        size = img.shape
        img1 = img
        gray2 = HED(img1, net)
        result.append(gray2)
    return result


# 找到与某个圆心所有相近的圆心, 被 identify_all 调用
def find_together_for_one(index: int, all: list):
    remains = [i for i in range(0, len(all))]
    remains.pop(index)
    together = [index]
    has_new = False
    while True:
        has_new = False
        for i in range(len(remains) - 1, -1, -1):
            this_x = all[remains[i]][0]
            this_y = all[remains[i]][1]
            for j in together:
                if (((this_x - all[j][0])**2 + (this_y - all[j][1])**2)**0.5 < ALLOWED_DISTANCE):
                    together.append(remains[i])
                    remains.pop(i)
                    has_new = True
                    break
        if (has_new == False):
            break
    result = []
    for i in together:
        result.append(all[i])
    return result


# 识别所有大圆圆心, 并进行挑选
def identify_all(all_img: list):
    result = []
    results = np.zeros((len(all_img), 2), np.float64)
    ratio = 1
    i = 1
    for img in all_img:
        # gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        gray = img
        # detect the center of bolts relative to the centre of
        (H, W) = gray.shape[:2]
        circleshed = cv2.HoughCircles(gray, cv2.HOUGH_GRADIENT, 1, round(H / 2),
                                      param1=100, param2=30,
                                      minRadius=round(H / 6), maxRadius=round(H / 2))
        center_x = 0
        center_y = 0
        all_circles = []
        for j in circleshed[0, :]:  # 遍历矩阵每一行的数据
            center_x = j[0]
            center_y = j[1]
            all_circles.append([center_x * ratio, center_y * ratio])
        closest_index = -1
        closest_distance = 10000000.0
        for k in range(0, len(all_circles)):
            distance = ((all_circles[k][0] - CENTER[0])**2 +
                        (all_circles[k][1] - CENTER[1])**2)**0.5
            if (distance < closest_distance):
                closest_index = k
                closest_distance = distance
        results[i - 1][0] = all_circles[closest_index][0]
        results[i - 1][1] = all_circles[closest_index][1]
        i += 1
    all_loc = []
    valid_locations = []
    for s in results:
        all_loc.append([float(s[0]), float(s[1])])
    for i in range(0, len(all_loc)):
        res = find_together_for_one(i, all_loc)
        if (len(res) >= 0.75 * len(all_loc)):
            valid_locations = res
            break
    return valid_locations


def get_small_chessboard_area(transformed_imgs: list):
    up_cells = CHESSBOARD_SIZE[1] / 2 - CHESSBOARD_SIZE[1] * 0.1 - CHESSBOARD_SIZE[4]
    left_cells = CHESSBOARD_SIZE[0] / 2 - CHESSBOARD_SIZE[0] * 0.1 - CHESSBOARD_SIZE[2]
    right_cells = CHESSBOARD_SIZE[0] / 2 - CHESSBOARD_SIZE[0] * 0.1 - CHESSBOARD_SIZE[3]
    one_cell_size = CENTER[0] * 2 / SIZE
    left = int(CENTER[0] - one_cell_size * left_cells)
    right = int(CENTER[0] + one_cell_size * right_cells)
    up = int(CENTER[0] - one_cell_size * up_cells)
    down = int(CENTER[0] + one_cell_size * up_cells)
    result = []
    for img in transformed_imgs:
        result.append(img[up:down, left:right])
    return (up, left, result)


def avg_center(all_centers: list):
    all_x = 0
    all_y = 0
    for i in range(0, len(all_centers)):
        all_x += all_centers[i][0]
        all_y += all_centers[i][1]
    return (float(all_x / len(all_centers)), float(all_y / len(all_centers)))


def get_small_corners(small_imgs: list):
    all_upleft_x = 0
    all_upleft_y = 0
    all_upright_x = 0
    all_upright_y = 0
    all_downleft_x = 0
    all_downleft_y = 0
    all_downright_x = 0
    all_downright_y = 0
    i = 0
    valid_num = 0
    for img in small_imgs:
        i += 1
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        r = cv2.findChessboardCorners(img, (11, 8), None)
        if (r[0] == False):
            print(f"第{i}张图片小棋盘格角点识别失败")
            continue
        r = cv2.cornerSubPix(img, r[1], (5, 4), (-1, -1), (cv2.TERM_CRITERIA_MAX_ITER | cv2.TERM_CRITERIA_EPS, 30, 0.001))
        valid_num += 1
        all_upleft_x += getCorner2(r, 'upleft')[0]
        all_upleft_y += getCorner2(r, 'upleft')[1]
        all_upright_x += getCorner2(r, 'upright')[0]
        all_upright_y += getCorner2(r, 'upright')[1]
        all_downleft_x += getCorner2(r, 'downleft')[0]
        all_downleft_y += getCorner2(r, 'downleft')[1]
        all_downright_x += getCorner2(r, 'downright')[0]
        all_downright_y += getCorner2(r, 'downright')[1]
    return (
        (all_upleft_x / valid_num, all_upleft_y / valid_num),
        (all_upright_x / valid_num, all_upright_y / valid_num),
        (all_downleft_x / valid_num, all_downleft_y / valid_num),
        (all_downright_x / valid_num, all_downright_y / valid_num)
    )


def to_small_cb_coordinate(small_corners: tuple, big_x: float, big_y: float):
    sc = small_corners
    # 长方形上面那条边
    vec1 = (sc[1][0] - sc[0][0], sc[1][1] - sc[0][1])
    # 长方形左边那条边
    vec2 = (sc[2][0] - sc[0][0], sc[2][1] - sc[0][1])
    # 大圆圆心位置
    vec_center = (big_x - sc[0][0], big_y - sc[0][1])
    a = (vec2[0] * vec_center[1] - vec2[1] * vec_center[0]) / (vec1[1] * vec2[0] - vec1[0] * vec2[1])
    b = (vec1[1] * vec_center[0] - vec_center[1] * vec1[0]) / (vec2[0] * vec1[1] - vec2[1] * vec1[0])
    return (10 * a, 7 * b)

if (__name__ == '__main__'):
     # 读取相机参数
     mtx, dist, newcameramtx = read_camera_matrix()
     # 读取所有图片
     all_imgs = read_all()
     # 去除畸变
     undistorted_imgs = undistort_all(all_imgs)
     # 透视变换
     transformed_imgs = transform_all(undistorted_imgs)
     # 进行 HED
     hedded_imgs = hed_all(transformed_imgs)
     # 识别大圆圆心
     all_centers = identify_all(hedded_imgs)
     if (len(all_centers) <= 0.6 * IMG_NUM):
         print(f'有效结果过少, 只有{len(all_centers)}个, 请检查原因')
     else:
         print(f'有效结果{len(all_centers)}个')
         # 将结果写入 Excel
         wb = Workbook()
         ws = wb.active
         for i in range(0, len(all_centers)):
             write_cell(ws, 1, i + 1, all_centers[i][0])
             write_cell(ws, 2, i + 1, all_centers[i][1])
         wb.save(OUTPUT)


def identify_center(camere_para_path: str, img_path: str, img_num: int, width: int, height: int, use_part_chessboard: bool, chessboard_size: tuple = None):
    global CAMERA_PARA_PATH, IMG_PATH, IMG_NUM, OUTPUT, WIDTH, HEIGHT, SIZE, USE_PART_CHESSBOARD, CHESSBOARD_SIZE
    CAMERA_PARA_PATH = camere_para_path
    IMG_NUM = img_num
    IMG_PATH = img_path
    # OUTPUT = output
    WIDTH = width - 1
    HEIGHT = height - 1
    SIZE = width + height
    USE_PART_CHESSBOARD = use_part_chessboard
    CHESSBOARD_SIZE = chessboard_size
    global mtx, dist, newcameramtx
    if (USE_PART_CHESSBOARD):
        separate_cb.CHESSBOARD_SIZE = CHESSBOARD_SIZE
    # 读取相机参数
    mtx, dist, newcameramtx = read_camera_matrix()
    # 读取所有图片
    all_imgs = read_all()
    print("图片读取成功")
    # 去除畸变
    undistorted_imgs = undistort_all(all_imgs)
    print("去除畸变完成")
    # 透视变换
    transformed_imgs = transform_all(undistorted_imgs)
    up_px, left_px, small_area = get_small_chessboard_area(transformed_imgs)
    print("透视变换完成")
    # 进行 HED
    hedded_imgs = hed_all(transformed_imgs)
    print("HED完成")
    # 识别大圆圆心
    all_centers = identify_all(hedded_imgs)
    if (len(all_centers) <= 0.6 * IMG_NUM):
        print(f'有效结果过少, 只有{len(all_centers)}个, 请检查原因')
        return (False, f'有效结果过少, 只有{len(all_centers)}个, 请检查原因')
    else:
        big_x, big_y = avg_center(all_centers)
        big_x -= left_px
        big_y -= up_px
        print(f'有效结果{len(all_centers)}个')
        return to_small_cb_coordinate(get_small_corners(small_area), big_x, big_y)
        # # 将结果写入 Excel
        # wb = Workbook()
        # ws = wb.active
        # for i in range(0, len(all_centers)):
        #     write_cell(ws, 1, i + 1, all_centers[i][0])
        #     write_cell(ws, 2, i + 1, all_centers[i][1])
        # wb.save(OUTPUT)
'''if (__name__ == '__main__'):
    identify_center(CAMERA_PARA_PATH, IMG_PATH, IMG_NUM, WIDTH, HEIGHT,  USE_PART_CHESSBOARD, CHESSBOARD_SIZE)'''