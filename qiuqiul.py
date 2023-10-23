CHESSBOARD_SIZE = (19, 19, 5, 5, 5)# 棋盘格尺寸: 5 个参数, 从左到右依次为 宽度、高度、左侧宽度、右侧宽度、上方高度 (尺寸为方格数量)
import glob
import numpy as np
import cv2
from sklearn.cluster import KMeans
import pandas as pd
import math
import openpyxl
from openpyxl import Workbook
import os
import separate_cb
separate_cb.CHESSBOARD_SIZE = CHESSBOARD_SIZE
from separate_cb import getFourCorners

graphtype=38        #参考小圆柱中心（范例0cm设定为38 (y=3)，范例3cm设定为16 (y=1)）
separate_cb.CHESSBOARD_SIZE = CHESSBOARD_SIZE
address=r'C:\Users\admin\Desktop\Codes-Zhengwei-20221219'
append1='\*.jpg'
append2='\{num}.jpg'
SOURCE_CALIBRATION = address+'\Calibration\small'+append1    #四小圆柱相机标定
SOURCE = address+ '/Users/zhengbaoqin/Desktop/project/1203/1203/s3cm'+append1   #四小圆柱中心计算用图片
RESULT = address+'\Temp1'+append2       #透视变换图片
SOURCE_PATH = address+'\Temp1'+append1
RESULT_PATH = address+'\Temp2'+ append2  #缩放后图片
SOURCE_PATHnew = address+ '\Temp2'+append1

PROTO_PATH = address+"\deploy.prototxt"
MODEL_PATH = address+"\hed_pretrained_bsds.caffemodel"

OUTPUTtmp = address+'\data.xlsx'
CAMERA_PARA_PATH = address+'\data.xlsx'
IMG_PATH = address+'\Large_3cm'          #大圆中心计算用图片
#参数
PATHcali = address+"\Calibration\large"  #大圆相机标定
IMG_NUM = 8# 照片数量
IMG_NUM2=16
WIDTH = 18# 棋盘格的宽度
HEIGHT = 18# 棋盘格的高度
SIZE = 40# 透视变换后的高度相当于多少个棋盘格
USE_PART_CHESSBOARD = True# 是否为部分棋盘格
CHESSBOARD_SIZE = (19, 19, 5, 5, 5)# 棋盘格尺寸: 5 个参数, 从左到右依次为 宽度、高度、左侧宽度、右侧宽度、上方高度 (尺寸为方格数量)
CENTER = (0.0, 0.0)
ALLOWED_DISTANCE = 8.0


CRITERIA = (cv2.TERM_CRITERIA_MAX_ITER | cv2.TERM_CRITERIA_EPS, 30, 0.001)
CHESSBOARD_SIZE = (11, 8)

class CropLayer(object):
    def __init__(self, params, blobs):
        # 初始化剪切区域开始和结束点的坐标
        self.xstart = 0
        self.ystart = 0
        self.xend = 0
        self.yend = 0

    # 计算输入图像的体积
    def getMemoryShapes(self, inputs):
        (inputShape, targetShape) = (inputs[0], inputs[1])
        (batchSize, numChannels) = (inputShape[0], inputShape[1])
        (H, W) = (targetShape[2], targetShape[3])

        # 计算开始和结束剪切坐标的值
        self.xstart = int((inputShape[3] - targetShape[3]) // 2)
        self.ystart = int((inputShape[2] - targetShape[2]) // 2)
        self.xend = self.xstart + W
        self.yend = self.ystart + H

        # 返回体积，接下来进行实际裁剪
        return [[batchSize, numChannels, H, W]]

    def forward(self, inputs):
        return [inputs[0][:, :, self.ystart:self.yend, self.xstart:self.xend]]

def HED(image, net, downW=800, downH=800):
    (H, W) = image.shape[:2]
    image = cv2.resize(image, (downW, downH))
    # 根据输入图像为全面的嵌套边缘检测器（Holistically-Nested Edge Detector）构建一个输出blob
    blob = cv2.dnn.blobFromImage(image, scalefactor=1.0, size=(800, 800),
                                mean=(104.00698793, 116.66876762, 122.67891434),
                                swapRB=False, crop=True)
    # # 设置blob作为网络的输入并执行算法以计算边缘图
    #print("[INFO] performing holistically-nested edge detection...")
    net.setInput(blob)
    hed = net.forward()
    # 调整输出为原始图像尺寸的大小
    hed = cv2.resize(hed[0, 0], (W, H))
    # 将图像像素缩回到范围[0,255]并确保类型为“UINT8”
    hed = (255 * hed).astype("uint8")
    return hed

def crop(image,cx,cy):
    size = image.shape
    gray=cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    ret_1, corners = cv2.findChessboardCorners(gray, (11, 8), None)
    criteria = (cv2.TERM_CRITERIA_MAX_ITER | cv2.TERM_CRITERIA_EPS, 30, 0.001)
    corners = cv2.cornerSubPix(gray, corners, (5, 5), (-1, -1), criteria)
    xcenter=corners[38][0][0]
    ycenter=corners[38][0][1]

    # y_start=math.ceil(ycenter-(size[0]-ycenter)/cy)
    # y_end=math.ceil(ycenter+(size[0]-ycenter)/cy)
    # x_start=math.ceil(xcenter-(size[1]-xcenter)/cx)
    # x_end=math.ceil(xcenter+(size[1]-xcenter)/cx)
    cropped = image[math.ceil(ycenter-(size[0]-ycenter)/cy):math.ceil(ycenter+(size[0]-ycenter)/cy),math.ceil(xcenter-(size[1]-xcenter)/cx):math.ceil(xcenter+(size[1]-xcenter)/cx)]
    # cv2.imwrite(RESULT_PATH.format(num = count+1), cropped)
    # print("Crop Finished")
    return cropped

def local_grid(gray, param):
    # detect the center of bolts relative to the centre of
    (H, W) = gray.shape[:2]
    circleshed = cv2.HoughCircles(gray, cv2.HOUGH_GRADIENT, 1, 20,
                                 param1=100, param2=param,
                                 minRadius=30, maxRadius=40)
    # use K-means algorithm to group the circles and extract the minimum from each groups
    '''
    km = KMeans(n_clusters=4).fit(circleshed[0, :, 0:2])
    # km = circleshed[0,:,0:2]
    df = pd.DataFrame({'X': circleshed[0, :, 0],
                       'Y': circleshed[0, :, 1],
                       'Radius': circleshed[0, :, 2],
                       'Label': km.labels_}
                      )
    mins = df.sort_values('Y', ascending=False).groupby('Label', as_index=False).first()
    '''
    df = pd.DataFrame({'X': circleshed[0, :, 0],
                       'Y': circleshed[0, :, 1],
                       'Radius': circleshed[0, :, 2]})
    df["Distance"]=(df["Y"]-H/2)**2+(df["X"]-W/2)**2
    df=df[df["Distance"]>=10000]
    mins = df.sort_values('Distance', ascending=False).tail(4)
    # print(mins)
    result = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)
    x_sum = 0
    y_sum = 0
    # show the circles
    for index, i in mins.iterrows():
        center = (i['X'].astype("int"), i['Y'].astype("int"))
        cv2.circle(result, center, 1, (255, 255, 255), 3)
        radius = i['Radius'].astype("int")
        cv2.circle(result, center, radius, (255, 255, 255), 3)
        # x_sum = x_sum + i['X'].astype("float64")
        # y_sum = y_sum + i['Y'].astype("float64")
        x_sum = x_sum + i['X']
        y_sum = y_sum + i['Y']
    x_sum = x_sum + 1
    # print('measured center: (', x_sum / 4, ",", y_sum / 4,")")

    return result,x_sum,y_sum

def detect_bolts(img, net):
    size = img.shape
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # 按真实长度缩放
    criteria = (cv2.TERM_CRITERIA_MAX_ITER | cv2.TERM_CRITERIA_EPS, 30, 0.001)
    ret_1, corners = cv2.findChessboardCorners(gray, (11, 8), None)
    corners = cv2.cornerSubPix(gray, corners, (5, 5), (-1, -1), criteria)
    
    center = corners[graphtype]
    # print('actual center: ', center)
    center_x = center[0][0]
    center_y = center[0][1]
    zeropoint = corners[0]
    center_xx = zeropoint[0][0]
    center_yy = zeropoint[0][1]
    #print('初始值： （',"%.2f"  % center_xx,",", "%.2f" % center_yy,")")
    #print("actual center: (", "%.2f" % center_x,",", "%.2f" % center_y,")")

    # 求螺栓中心坐标
    gray2 = HED(img,net)
    result,x_sum,y_sum = local_grid(gray2, 18)
    x_diff = x_sum / 4 - center_x
    y_diff = y_sum / 4 - center_y
    aa=(x_sum/4-center_xx)/15
    bb=(y_sum/4-center_yy)/15
    print('四小圆柱计算出的中心:', aa,' ',bb)
    a=(center_x-center_xx)/15
    b=(center_y-center_yy)/15
    print('认为的中心:', a,' ',b)
    diff = math.sqrt(x_diff * x_diff + y_diff * y_diff)
    # difference between measured center and actual center
    # print(diff)
    # print('difference: ')
    # print("%.2f" % diff)
    return a,b,aa,bb,result,diff


# function to detect the outside base circle with HoughCircle and HED
# return the circle center
def detect_direct_hough_with_hed(image, net):
    hed = HED(image, net)
    (H, W) = hed.shape[:2]
    circleshed = cv2.HoughCircles(hed, cv2.HOUGH_GRADIENT, 1, round(H / 2),
                                 param1=100, param2=60,
                                 minRadius=round(H / 6), maxRadius=round(H / 2))
    center_x = 0
    center_y = 0
    for i in circleshed[0, :]:  # 遍历矩阵每一行的数据
        cv2.circle(image, (int(i[0]), int(i[1])), int(i[2]), (0, 255, 0), 2)
        cv2.circle(image, (int(i[0]), int(i[1])), 2, (0, 0, 255), 3)
        # above is to draw the circle and circle center on image, if you would like to check, you could use imshow
        center_x = i[0]
        center_y = i[1]
    return center_x, center_y
# 获取每张照片的路径

def getPathcali(n: int):
    n = str(n)
    if (len(n) == 1):
        n = '0' + n
    return f'{PATHcali}/{n}.jpg'

# function to call detect_direct_hough_with_hed
def detect_base():
    images=glob.glob(SOURCE_PATH)
    for image in enumerate(images):
        img = cv2.imread(image)
        x, y = img.shape[0:2]
        img = cv2.resize(img, (int(y / 6), int(x / 6)))
        x_center, y_center = detect_direct_hough_with_hed(img, net)
        print("x_center: ", x_center)
        print("y_center: ", y_center)
        return x_center,y_center
# recover the perspective of the picture, i.e. make the chessborad in the z=0 surface.
def perspective_recover(image, criteria, w, h, size=4000, ratio=0.03):
    
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # Find the chess board corners
    ret_1, corners = cv2.findChessboardCorners(gray, (11, 8), None)
    corners = cv2.cornerSubPix(gray, corners, (5, 5), (-1, -1), criteria)
    # prepare corresponding points for homography
    objp2 = np.zeros((11 * 8, 2), np.float32)
    cap = max(h, w)
    k = np.mgrid[-5:6:1, -4:4:1].T.reshape(-1, 2)
    objp2[:, :2] = (k * ratio + 0.5) * cap
    pts1 = np.float32([corners[0][0], corners[10][0], corners[-1][0], corners[-11][0]])
    # rectify picture to proper orientation that make the bolts face downward
    vec = corners[10][0] - corners[0][0]
    tan = vec[1] / vec[0]
    if tan <= 1 and tan >= -1:
        if vec[0] >= 0:  # 0 10 -1 -11
            pts2 = np.float32([objp2[0], objp2[10], objp2[-1], objp2[-11]])
        else:
            pts2 = np.float32([objp2[-1], objp2[-11], objp2[0], objp2[10]])
    else:
        if vec[1] >= 0:
            pts2 = np.float32([objp2[10], objp2[-1], objp2[-11], objp2[0]])
        else:
            pts2 = np.float32([objp2[-11], objp2[0], objp2[10], objp2[-1]])
    # obtain and perform transformation
    M = cv2.getPerspectiveTransform(pts1, pts2)
    pts = np.float32([[0, 0], [0, h], [w, h], [w, 0]]).reshape(-1, 1, 2)
    pts_ = cv2.perspectiveTransform(pts, M)
    [x_min, y_min] = np.int32(pts_.min(axis=0).ravel()-0.5)
    [x_max, y_max] = np.int32(pts_.max(axis=0).ravel()+0.5)
    diff = [-x_min, -y_min]
    H_diff = np.array([[1, 0, diff[0]], [0, 1, diff[1]], [0, 0, 1]])
    H = H_diff.dot(M)
    dst = cv2.warpPerspective(image, H, (x_max-x_min, y_max-y_min))
    return dst

# 先声明需要使用的全局变量
mtx = None
dist = None
newcameramtx = None


def getPath(n: int):
    n = str(n)
    if (len(n) == 1):
        n = '0' + n
    return f'{IMG_PATH}/{n}.jpg'
# Excel 部分
ALPHAS = '.ABCDEFGHIJKLMNOPQRSTUVWXYZ'
NUMS = {}
for i in range(1, 27):
    NUMS[ALPHAS[i]] = i


def numToAlpha(n):
    s = ''
    while (n > 0):
        t = n % 26
        if (t == 0):
            t = 26
        s = ALPHAS[t] + s
        n = (n - t) // 26
    return s
def alphaToNum(a):
    l = len(a)
    sum = 0
    for i in range(0, l):
        sum += 26**(l - i - 1) * NUMS[a[i]]
    return sum
def cellToTuple(cell: str):
    for i in range(0, len(cell)):
        try:
            a = NUMS[cell[i]]
        except:
            n = i
            break
    return (alphaToNum(cell[:n]), int(cell[n:]))
def tupleToCell(t):
    return numToAlpha(t[0]) + str(t[1])
def read_cell(sheet, i, j):
    return sheet[tupleToCell((i, j))].value


def write_cell(sheet, i, j, value):
    sheet[tupleToCell((i, j))].value = value


# 从 Excel 读取相机参数的函数
def read_camera_matrix():
    mtx = np.zeros((3, 3), np.float64)
    newcameramtx = np.zeros((3, 3), np.float64)
    dist = np.zeros((1, 5), np.float64)
    wb = openpyxl.load_workbook(CAMERA_PARA_PATH)
    ws = wb.active
    mtx[0][0] = read_cell(ws, 1, 1)
    mtx[0][1] = read_cell(ws, 2, 1)
    mtx[0][2] = read_cell(ws, 3, 1)
    mtx[1][0] = read_cell(ws, 1, 2)
    mtx[1][1] = read_cell(ws, 2, 2)
    mtx[1][2] = read_cell(ws, 3, 2)
    mtx[2][0] = read_cell(ws, 1, 3)
    mtx[2][1] = read_cell(ws, 2, 3)
    mtx[2][2] = read_cell(ws, 3, 3)
    newcameramtx[0][0] = read_cell(ws, 1, 5)
    newcameramtx[0][1] = read_cell(ws, 2, 5)
    newcameramtx[0][2] = read_cell(ws, 3, 5)
    newcameramtx[1][0] = read_cell(ws, 1, 6)
    newcameramtx[1][1] = read_cell(ws, 2, 6)
    newcameramtx[1][2] = read_cell(ws, 3, 6)
    newcameramtx[2][0] = read_cell(ws, 1, 7)
    newcameramtx[2][1] = read_cell(ws, 2, 7)
    newcameramtx[2][2] = read_cell(ws, 3, 7)
    dist[0][0] = read_cell(ws, 1, 9)
    dist[0][1] = read_cell(ws, 2, 9)
    dist[0][2] = read_cell(ws, 3, 9)
    dist[0][3] = read_cell(ws, 4, 9)
    dist[0][4] = read_cell(ws, 5, 9)
    return (mtx, dist, newcameramtx)
# 读取所有图片
def read_all():
    all = []
    for i in range(1, IMG_NUM + 1):
        all.append(cv2.imread(getPath(i)))
    return all
# 对所有图片去除畸变
def undistort_all(all_img: list):
    result = []
    for img in all_img:
        dst = cv2.undistort(img, mtx, dist, None, newcameramtx)
        result.append(dst)
    return result
# 在棋盘格所有格点中挑选出角上的, 被 transform_all 调用
def getCorner2(points: np.ndarray, Type: str):
    # type  upleft, upright, downleft, downright
    dic = {
        'upleft': [1, 1],
        'upright': [-1, 1],
        'downleft': [1, -1],
        'downright': [-1, -1]
    }
    mode = dic[Type]
    sum = []
    for p in points:
        sum.append(p[0][0] * mode[0] + p[0][1] * mode[1])
    loc = sum.index(min(sum))
    return points[loc][0]
# 对所有图片透视变换
def transform_all(all_img: list):
    global CENTER
    result = []
    if (USE_PART_CHESSBOARD):
        for img in all_img:
            cv2.imwrite('./._tmp_.png', img)
            fourCorners = getFourCorners('./._tmp_.png')
            os.remove('./._tmp_.png')
            if (fourCorners[0] == False):
                print(fourCorners[1])
                continue
            upleft = fourCorners[1][0]
            upright = fourCorners[1][1]
            downleft = fourCorners[1][2]
            downright = fourCorners[1][3]
            h, w = img.shape[:2]
            cap = max(h, w)
            cap = h
            CENTER = (float(0.5 * cap), float(0.5 * cap))
            x_size = SIZE
            objPoints = np.float32([[x_size / 2 - WIDTH / 2 + 0.5, SIZE / 2 - HEIGHT / 2 + 0.5], [x_size / 2 + WIDTH / 2 - 0.5, SIZE / 2 - HEIGHT / 2 + 0.5], [x_size / 2 - WIDTH / 2 + 0.5, SIZE / 2 + HEIGHT / 2 - 0.5], [x_size / 2 + WIDTH / 2 - 0.5, SIZE / 2 + HEIGHT / 2 - 0.5]]) * cap / SIZE
            imgPoints = np.float32([upleft, upright, downleft, downright])
            m = cv2.getPerspectiveTransform(imgPoints, objPoints)
            new = cv2.warpPerspective(img, m, (cap, cap))
            result.append(new)
        return result
    i = 0
    for img in all_img:
        i += 1
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        h, w = img.shape[:2]
        cap = h
        CENTER = (float(0.5 * cap), float(0.5 * cap))
        x_size = SIZE
        objPoints = np.float32(
            [[x_size / 2 - WIDTH / 2 + 0.5, SIZE / 2 - HEIGHT / 2 + 0.5], [x_size / 2 + WIDTH / 2 - 0.5, SIZE / 2 - HEIGHT / 2 + 0.5], [x_size / 2 - WIDTH / 2 + 0.5, SIZE / 2 + HEIGHT / 2 - 0.5], [x_size / 2 + WIDTH / 2 - 0.5, SIZE / 2 + HEIGHT / 2 - 0.5]]) * cap / SIZE
        corners = cv2.findChessboardCorners(gray, (WIDTH, HEIGHT), None)
        if (corners[0] == False):
            print(f'第{i}张图片透视变换是未成功识别棋盘格, 请检查原因')
            continue
        corners = corners[1]
        corners = cv2.cornerSubPix(gray, corners, (5, 4), (-1, -1),
                                   (cv2.TERM_CRITERIA_MAX_ITER | cv2.TERM_CRITERIA_EPS, 30, 0.001))
        upleft = getCorner2(corners, 'upleft')
        upright = getCorner2(corners, 'upright')
        downleft = getCorner2(corners, 'downleft')
        downright = getCorner2(corners, 'downright')
        imgPoints = np.float32([upleft, upright, downleft, downright])
        m = cv2.getPerspectiveTransform(imgPoints, objPoints)
        new = cv2.warpPerspective(img, m, (cap, cap))
        result.append(new)
    return result
# 对所有图片进行 HED
def hed_all(all_img: list):
    result = []
    net = cv2.dnn.readNetFromCaffe(PROTO_PATH, MODEL_PATH)
    for img in all_img:
        size = img.shape
        ratio = 1
        ratiox = int(size[1] / ratio)
        ratioy = int(size[0] / ratio)
        img = cv2.resize(img, (ratiox, ratioy))
        size = img.shape
        img1 = img
        #net=None
        gray2 = HED(img1, net)
        result.append(gray2)
    return result
# 找到与某个圆心所有相近的圆心, 被 identify_all 调用
def find_together_for_one(index: int, all: list):
    remains = [i for i in range(0, len(all))]
    remains.pop(index)
    together = [index]
    has_new = False
    while True:
        has_new = False
        for i in range(len(remains) - 1, -1, -1):
            this_x = all[remains[i]][0]
            this_y = all[remains[i]][1]
            for j in together:
                if (((this_x - all[j][0])**2 + (this_y - all[j][1])**2)**0.5 < ALLOWED_DISTANCE):
                    together.append(remains[i])
                    remains.pop(i)
                    has_new = True
                    break
        if (has_new == False):
            break
    result = []
    for i in together:
        result.append(all[i])
    return result
# 识别所有大圆圆心, 并进行挑选
def identify_all(all_img: list):
    result = []
    results = np.zeros((len(all_img), 2), np.float64)
    ratio = 1
    i = 1
    for img in all_img:
        # gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        gray = img
        # detect the center of bolts relative to the centre of
        (H, W) = gray.shape[:2]
        circleshed = cv2.HoughCircles(gray, cv2.HOUGH_GRADIENT, 1, round(H / 2),
                                      param1=100, param2=30,
                                      minRadius=round(H / 6), maxRadius=round(H / 2))
        center_x = 0
        center_y = 0
        all_circles = []
        for j in circleshed[0, :]:  # 遍历矩阵每一行的数据
            center_x = j[0]
            center_y = j[1]
            all_circles.append([center_x * ratio, center_y * ratio])
        closest_index = -1
        closest_distance = 10000000.0
        for k in range(0, len(all_circles)):
            distance = ((all_circles[k][0] - CENTER[0])**2 +
                        (all_circles[k][1] - CENTER[1])**2)**0.5
            if (distance < closest_distance):
                closest_index = k
                closest_distance = distance
        results[i - 1][0] = all_circles[closest_index][0]
        results[i - 1][1] = all_circles[closest_index][1]
        i += 1
    all_loc = []
    valid_locations = []
    for s in results:
        all_loc.append([float(s[0]), float(s[1])])
    for i in range(0, len(all_loc)):
        res = find_together_for_one(i, all_loc)
        if (len(res) >= 0.75 * len(all_loc)):
            valid_locations = res
            break
    return valid_locations
def get_small_chessboard_area(transformed_imgs: list):
    up_cells = CHESSBOARD_SIZE[1] / 2 - CHESSBOARD_SIZE[1] * 0.1 - CHESSBOARD_SIZE[4]
    left_cells = CHESSBOARD_SIZE[0] / 2 - CHESSBOARD_SIZE[0] * 0.1 - CHESSBOARD_SIZE[2]
    right_cells = CHESSBOARD_SIZE[0] / 2 - CHESSBOARD_SIZE[0] * 0.1 - CHESSBOARD_SIZE[3]
    one_cell_size = CENTER[0] * 2 / SIZE
    left = int(CENTER[0] - one_cell_size * left_cells)
    right = int(CENTER[0] + one_cell_size * right_cells)
    up = int(CENTER[0] - one_cell_size * up_cells)
    down = int(CENTER[0] + one_cell_size * up_cells)
    result = []
    for img in transformed_imgs:
        result.append(img[up:down, left:right])
    return (up, left, result)
def avg_center(all_centers: list):
    all_x = 0
    all_y = 0
    for i in range(0, len(all_centers)):
        all_x += all_centers[i][0]
        all_y += all_centers[i][1]
    return (float(all_x / len(all_centers)), float(all_y / len(all_centers)))
def get_small_corners(small_imgs: list):
    all_upleft_x = 0
    all_upleft_y = 0
    all_upright_x = 0
    all_upright_y = 0
    all_downleft_x = 0
    all_downleft_y = 0
    all_downright_x = 0
    all_downright_y = 0
    i = 0
    valid_num = 0
    for img in small_imgs:
        i += 1
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        r = cv2.findChessboardCorners(img, (11, 8), None)
        if (r[0] == False):
            print(f"第{i}张图片小棋盘格角点识别失败")
            continue
        r = cv2.cornerSubPix(img, r[1], (5, 4), (-1, -1), (cv2.TERM_CRITERIA_MAX_ITER | cv2.TERM_CRITERIA_EPS, 30, 0.001))
        valid_num += 1
        all_upleft_x += getCorner2(r, 'upleft')[0]
        all_upleft_y += getCorner2(r, 'upleft')[1]
        all_upright_x += getCorner2(r, 'upright')[0]
        all_upright_y += getCorner2(r, 'upright')[1]
        all_downleft_x += getCorner2(r, 'downleft')[0]
        all_downleft_y += getCorner2(r, 'downleft')[1]
        all_downright_x += getCorner2(r, 'downright')[0]
        all_downright_y += getCorner2(r, 'downright')[1]
    return (
        (all_upleft_x / valid_num, all_upleft_y / valid_num),
        (all_upright_x / valid_num, all_upright_y / valid_num),
        (all_downleft_x / valid_num, all_downleft_y / valid_num),
        (all_downright_x / valid_num, all_downright_y / valid_num)
    )

def to_small_cb_coordinate(small_corners: tuple, big_x: float, big_y: float):
    sc = small_corners
    # 长方形上面那条边
    vec1 = (sc[1][0] - sc[0][0], sc[1][1] - sc[0][1])
    # 长方形左边那条边
    vec2 = (sc[2][0] - sc[0][0], sc[2][1] - sc[0][1])
    # 大圆圆心位置
    vec_center = (big_x - sc[0][0], big_y - sc[0][1])
    a = (vec2[0] * vec_center[1] - vec2[1] * vec_center[0]) / (vec1[1] * vec2[0] - vec1[0] * vec2[1])
    b = (vec1[1] * vec_center[0] - vec_center[1] * vec1[0]) / (vec2[0] * vec1[1] - vec2[1] * vec1[0])
    return (10 * a, 7 * b)

def write_camera_matrix(mtx, dist, newcameramtx):
    wb = Workbook()
    ws = wb.active
    write_cell(ws, 1, 1, float(mtx[0][0]))
    write_cell(ws, 2, 1, float(mtx[0][1]))
    write_cell(ws, 3, 1, float(mtx[0][2]))
    write_cell(ws, 1, 2, float(mtx[1][0]))
    write_cell(ws, 2, 2, float(mtx[1][1]))
    write_cell(ws, 3, 2, float(mtx[1][2]))
    write_cell(ws, 1, 3, float(mtx[2][0]))
    write_cell(ws, 2, 3, float(mtx[2][1]))
    write_cell(ws, 3, 3, float(mtx[2][2]))
    write_cell(ws, 1, 5, float(newcameramtx[0][0]))
    write_cell(ws, 2, 5, float(newcameramtx[0][1]))
    write_cell(ws, 3, 5, float(newcameramtx[0][2]))
    write_cell(ws, 1, 6, float(newcameramtx[1][0]))
    write_cell(ws, 2, 6, float(newcameramtx[1][1]))
    write_cell(ws, 3, 6, float(newcameramtx[1][2]))
    write_cell(ws, 1, 7, float(newcameramtx[2][0]))
    write_cell(ws, 2, 7, float(newcameramtx[2][1]))
    write_cell(ws, 3, 7, float(newcameramtx[2][2]))
    write_cell(ws, 1, 9, float(dist[0][0]))
    write_cell(ws, 2, 9, float(dist[0][1]))
    write_cell(ws, 3, 9, float(dist[0][2]))
    write_cell(ws, 4, 9, float(dist[0][3]))
    write_cell(ws, 5, 9, float(dist[0][4]))
    wb.save(OUTPUTtmp)

def identify_center(camere_para_path: str, img_path: str, img_num: int, width: int, height: int, use_part_chessboard: bool, chessboard_size: tuple = None):
    global CAMERA_PARA_PATH, IMG_PATH, IMG_NUM, OUTPUT, WIDTH, HEIGHT, SIZE, USE_PART_CHESSBOARD, CHESSBOARD_SIZE
    CAMERA_PARA_PATH = camere_para_path
    IMG_NUM = img_num
    IMG_PATH = img_path
    # OUTPUT = output
    WIDTH = width - 1
    HEIGHT = height - 1
    SIZE = width + height
    USE_PART_CHESSBOARD = use_part_chessboard
    CHESSBOARD_SIZE = chessboard_size
    global mtx, dist, newcameramtx
    if (USE_PART_CHESSBOARD):
        separate_cb.CHESSBOARD_SIZE = CHESSBOARD_SIZE
    # 读取相机参数
    mtx, dist, newcameramtx = read_camera_matrix()
    # 读取所有图片
    all_imgs = read_all()
    print("图片读取成功")
    # 去除畸变
    undistorted_imgs = undistort_all(all_imgs)
    print("去除畸变完成")
    # 透视变换
    transformed_imgs = transform_all(undistorted_imgs)
    up_px, left_px, small_area = get_small_chessboard_area(transformed_imgs)
    print("透视变换完成")
    # 进行 HED
    hedded_imgs = hed_all(transformed_imgs)
    print("HED完成")
    # 识别大圆圆心
    all_centers = identify_all(hedded_imgs)
    if (len(all_centers) <= 0.6 * IMG_NUM):
        print(f'有效结果过少, 只有{len(all_centers)}个, 请检查原因')
        return (False, f'有效结果过少, 只有{len(all_centers)}个, 请检查原因')
    else:
        big_x, big_y = avg_center(all_centers)
        big_x -= left_px
        big_y -= up_px
        print(f'有效结果{len(all_centers)}个')
        return to_small_cb_coordinate(get_small_corners(small_area), big_x, big_y)
        # # 将结果写入 Excel
if __name__ == '__main__':
    
    #圆柱
    # read grayscale images from for_calibration
    calibration_img = [cv2.imread(file, 0) for file in glob.glob(SOURCE_CALIBRATION)]
    objp= np.zeros((11 * 8, 3), np.float32)
    objp[:, :2] = np.mgrid[0:11, 0:8].T.reshape(-1, 2)
    img_points = []
    obj_points = []


    for gray_image in calibration_img:
        # find and optimize chessboard corners
        retval, corners = cv2.findChessboardCorners(gray_image, CHESSBOARD_SIZE, flags=cv2.CALIB_CB_ADAPTIVE_THRESH +
                                                                      cv2.CALIB_CB_FAST_CHECK +
                                                                      cv2.CALIB_CB_NORMALIZE_IMAGE)
        corners2 = cv2.cornerSubPix(gray_image, corners, (5, 5), (-1, -1), CRITERIA)
        if [corners2]:
            img_points.append(corners2)
        else:
            img_points.append(corners)
        obj_points.append(objp)

        # camera calibration
        size_all = calibration_img[0].shape[::-1]
        retval, mtx, dist, rvec, tvec = cv2.calibrateCamera(obj_points, img_points, size_all, None, None)
        
        # optimize camera calibration
        h, w = calibration_img[0].shape[:2]
        new_camera_mtx, roi = cv2.getOptimalNewCameraMatrix(mtx, dist, (w, h), 1, (w, h))
    # read images for recognition
    images = [cv2.imread(file) for file in glob.glob(SOURCE)]
    undistorted_images = []
    for image in images:
        # undistort images
        undistorted = cv2.undistort(image, mtx, dist, None, new_camera_mtx)
        undistorted_images.append(undistorted)
    # perspective transform
    for count, undistorted_image in enumerate(undistorted_images):
        perspective = perspective_recover(undistorted_image, CRITERIA, w, h)
        cv2.imwrite(RESULT.format(num = count+1), perspective)
        
    #
    cv2.dnn_registerLayer("Crop", CropLayer)
    net = cv2.dnn.readNetFromCaffe(PROTO_PATH, MODEL_PATH)
    images.clear()
    images = [cv2.imread(file) for file in glob.glob(SOURCE_PATH)]
    for count, image in enumerate(images):
        height=image.shape[0]
        width=image.shape[1]
        size = image.shape
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        criteria = (cv2.TERM_CRITERIA_MAX_ITER | cv2.TERM_CRITERIA_EPS, 30, 0.001)
        ret_1, corners = cv2.findChessboardCorners(gray, (11, 8), None)
        corners = cv2.cornerSubPix(gray, corners, (5, 5), (-1, -1), criteria)
        # print(corners[44][0])
        pixel = corners[12][0] - corners[13][0]
        d = math.sqrt(pixel[0] * pixel[0] + pixel[1] * pixel[1])
        ratio = d / 15
        # print(ratio)
        ratiox = int(size[1] / ratio)
        ratioy = int(size[0] / ratio)
        image = cv2.resize(image, (ratiox, ratioy))

        cv2.imwrite(RESULT_PATH.format(num = count+1), image) 
    net = cv2.dnn.readNetFromCaffe(PROTO_PATH, MODEL_PATH)
    images = [cv2.imread(file) for file in glob.glob(SOURCE_PATHnew)]
    difference = []
    amean=[];
    bmean=[];
    aamean=[];
    bbmean=[];
    for count, image in enumerate(images):
        cx=2
        cy=2
        cropped=crop(image,cx,cy)
        # cv2.imwrite(RESULT_PATH.format(num = count+1), cropped)
        # result = detect_bolts(image,net)
        a,b,aa,bb,result,diff = detect_bolts(cropped, net)
        if(aa>0 and bb>0):
            print('误差：',diff)
            difference.append(diff)
            aamean.append(aa)
            bbmean.append(bb)
            amean.append(a)
            bmean.append(b)
        # print(diff)
        # cv2.imwrite(RESULT_PATH.format(num = count+1), result)
    difference = np.array(difference)
    aamean=np.array(aamean)
    bbmean=np.array(bbmean)
    aam=np.mean(aamean)
    bbm=np.mean(bbmean)
    am=np.mean(amean)
    bm=np.mean(bmean)
    print('小圆柱计算中心坐标(单位：网格):[',aam,',',bbm,']')
    print('小圆柱参考中心坐标(单位：网格):[', am, ',',bm,']')
    print('小圆柱中心误差值(单位：mm):',np.mean(difference))
    print('小圆柱中心误差值方差(mm**2):',np.var(difference))
    # print(type(difference))
    
    #大圆部分
    # 读取相机参数
    print('big circle:')
    
    
    corners = []
    objPoint = np.zeros((WIDTH * HEIGHT, 3), np.float32)
    objPoint[:, :2] = np.mgrid[0:WIDTH, 0:HEIGHT].T.reshape(-1, 2)
    objPoints = []

    # 棋盘格角点识别
    for i in range(1, IMG_NUM2 + 1):
        print(f"第{i}张照片开始识别")
        img = cv2.imread(getPathcali(i))
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        ret, corner = cv2.findChessboardCorners(gray, (WIDTH, HEIGHT), None)
        # corner 有可能返回 None, 但不能直接用 ==None 判断
        try:
            if (len(corner) != WIDTH * HEIGHT):
                print(f"第{i}张照片未成功识别, 请检查原因")
                continue
            else:
                print(f"第{i}张照片识别成功")
        except:
            print(f"第{i}张照片未成功识别, 请检查原因")
            continue
        corner = cv2.cornerSubPix(gray, corner, (5, 4), (-1, -1),
                                  (cv2.TERM_CRITERIA_MAX_ITER | cv2.TERM_CRITERIA_EPS, 30, 0.001))
        corners.append(corner)
        objPoints.append(objPoint)

    # 相机标定
    size_all = gray.shape[::-1]
    ret, mtx, dist, rvecs, tvecs = cv2.calibrateCamera(
        objPoints, corners, size_all, None, None)

    # 相机参数优化
    dist = dist[0:4]
    h, w = gray.shape[:2]
    newcameramtx, roi = cv2.getOptimalNewCameraMatrix(mtx, dist, (w, h), 1, (w, h))

    print(mtx, dist, newcameramtx)

    
    write_camera_matrix(mtx, dist, newcameramtx)
    
    '''mtx, dist, newcameramtx = read_camera_matrix()
    # 读取所有图片
    all_imgs = read_all()
    # 去除畸变
    undistorted_imgs = undistort_all(all_imgs)
    # 透视变换
    transformed_imgs = transform_all(undistorted_imgs)
    # 进行 HED
    hedded_imgs = hed_all(transformed_imgs)
    # 识别大圆圆心
    all_centers = identify_all(hedded_imgs)
    if (len(all_centers) <= 0.6 * IMG_NUM):
         print(f'有效结果过少, 只有{len(all_centers)}个, 请检查原因')
    else:
         print(f'有效结果{len(all_centers)}个')
         #return to_small_cb_coordinate(get_small_corners(small_area), big_x, big_y)
         # 将结果写入 Excel
         wb = Workbook()
         ws = wb.active
         for i in range(0, len(all_centers)):
             write_cell(ws, 1, i + 1, all_centers[i][0])
             write_cell(ws, 2, i + 1, all_centers[i][1])
             print(all_centers[i][0],' ',all_centers[i][1])
         wb.save(OUTPUT)'''

    a,b=identify_center(CAMERA_PARA_PATH, 
                IMG_PATH, 8, 32, 32, True, (32, 32, 6, 6, 6))
    diffa=a-aam
    diffb=b-bbm
    print('大圆计算中心坐标(单位网格):[',a,',',b,']')
    x=graphtype%11
    y=graphtype/11
    print('大圆参考中心坐标(单位网格):[','5',',','3',']')
    diff=math.sqrt(((5-a)*15)**2+((3-b)*15)**2)
    print('大圆中心误差值(单位：mm):', diff)
    print(' ')
    
    diff=math.sqrt((diffa*15)**2+(diffb*15)**2)
    print('计算获得的两中心偏移量(单位：mm):', diff)
    diff2=math.sqrt(((5-am)*15)**2+((3-bm)*15)**2)
    #diff2=math.sqrt((x-5)**2+(y-3)**2)
    print('参考的两中心偏移量(单位：mm):',diff2)
    diff=diff-diff2
    print('偏移量误差(单位：mm):', diff)
    print('Finished.')


